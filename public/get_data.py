from openpyxl import load_workbook
import os


RegTelPath = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))+ '/test_data/test_case.xlsx'



class GetData:
    RegTel = load_workbook(RegTelPath)['init'].cell(1,1).value
    TOKEN = None
    tel_1 = '1211313131'


