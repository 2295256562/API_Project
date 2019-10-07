import ast

from openpyxl import load_workbook
from public.get_data import GetData
from public.do_regx import DoRegx

class DoExcel(object):
    """读取excel"""

    def __init__(self, file_name, sheet_name):
        """
        :param file_name: 文件名称
        :param sheet_name: sheet名称
        """
        self.file = file_name
        self.sheet = sheet_name
        self.files = load_workbook(self.file)
        self.sheets = self.files[self.sheet]

    def get_headers(self):
        """
        拿到excel头部
        :return: excel第一行的数据
        """
        headers = []
        for i in range(1, self.sheets.max_column + 1):
            headers.append(self.sheets.cell(1, i).value)
        return headers

    def get_data(self):
        """
        拿到excel中的数据
        :return:  返回excel中的数据
        """
        header = self.get_headers()
        test_data = []

        tel = getattr(GetData, 'RegTel')
        for i in range(2, self.sheets.max_row + 1):
            sub = {}
            # print(self.sheets.cell(i, 4).value.find('$tel_1'))
            for j in range(1, self.sheets.max_column + 1):
                sub[header[j-1]] = self.sheets.cell(i, j).value
            test_data.append(sub)

        # 处理data里的数据
        # for i in test_data:
        #     if i['data'].find('$tel_1') != -1:
        #         tel = i['data'].replace('$tel_1', str(tel))
        #         i['data'] = tel

        for i in test_data:
            tel = DoRegx.do_regx(i['data'])
            i['data'] = tel
        return test_data

    def write_back(self, i, result, TestResult):
        """
        写入数据到excel
        :param i: 行号
        :param j: 列
        :param result: res.json返回的结果
        :param TestResult: 是否通过  pass or fail
        :return:
        """
        self.sheets.cell(i, 7).value = result
        self.sheets.cell(i, 8).value = TestResult
        self.files.save(self.file)


    def update_tel(self, tel):
        """
        更新手机号
        :param tel: 手机号
        :return:
        """
        self.sheets.cell(2, 1).value = tel
        self.files.save(self.file)


if __name__ == '__main__':
    test_data = DoExcel(file_name=r'D:\Python代码\API_Project\test_data\xxxc.xlsx', sheet_name='112')
    data = test_data.get_data()
    print(data)
